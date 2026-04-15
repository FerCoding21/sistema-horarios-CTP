import openpyxl, re, glob
from collections import defaultdict

ruta = glob.glob(r'C:\Users\Fer\Downloads\HORARIO*.xlsx')[0]
wb = openpyxl.load_workbook(ruta, data_only=True)
GRUPOS = [s for s in wb.sheetnames if re.match(r'\d{2}-\d+', s)]
MAT_COLS = [3, 5, 7, 9, 11]

NORM = {
    # Académicas comunes
    'matemática': 'Matemática', 'matematica': 'Matemática', 'matematicas': 'Matemática',
    'español': 'Español',
    'inglés': 'Inglés', 'ingles': 'Inglés', 'ingles.': 'Inglés',
    'inglis for comunication': 'English For Communication', 'inglis for': 'English For Communication',
    'educ. física': 'Educación Física', 'educacion fisica': 'Educación Física',
    'educación fisica': 'Educación Física',
    'educ. física ana yency garro díaz': 'Educación Física',
    'educaci\u00f3n fisica': 'Educación Física',
    'est. sociales': 'Estudios Sociales', 'estudios sociales': 'Estudios Sociales',
    'estudios socilaes': 'Estudios Sociales', 'estudios soclaes': 'Estudios Sociales',
    'cívica': 'Cívica', 'civica': 'Cívica',
    'religión': 'Religión', 'religion': 'Religión',
    'música': 'Música', 'musica': 'Música',
    'física': 'Física', 'fisica': 'Física', 'fisica.': 'Física', 'fisica ': 'Física',
    'biología': 'Biología', 'biologia': 'Biología',
    'química': 'Química', 'quimica': 'Química',
    'psicología': 'Psicología', 'psicologia': 'Psicología',
    'emprendedurismo e innovación': 'Emprendedurismo',
    'emprendimiento e innovación aplicada a la csrcso': 'Emprendedurismo',
    'emprendimiento e innovación aplicada al secretariado ejecutivo': 'Emprendedurismo',
    'english for communication': 'English For Communication',
    # MP 10°
    'diseño y manufactura asistida por computadora': 'Diseño y Manufactura Asistida por Computadora',
    'mecanizado con máquinas herramientas': 'Mecanizado con Máquinas Herramientas',
    'operaciones de equipo de banco y metrología dimensional': 'Operaciones de Equipo de Banco y Metrología Dimensional',
    'tecnología de la información aplicada a la mp': 'Tecnología de la Información Aplicada a la MP',
    'dibujo lineal': 'Dibujo Lineal',
    'dibujo técnico asistido por computadora': 'Dibujo Técnico Asistido por Computadora',
    'english oriented to precision mechanics': 'English Oriented to Precision Mechanics',
    # MP 11°
    'dibujo asistido por computadora': 'Dibujo Técnico Asistido por Computadora',
    'dibijo mecanico': 'Dibujo Mecánico', 'dibujo mecanico': 'Dibujo Mecánico',
    'diseño mano factura asistida': 'Diseño y Manufactura Asistida por Computadora',
    'mecanizado maquinas herramientas': 'Mecanizado con Máquinas Herramientas',
    'tics': 'TICs',
    # MP 12°
    'dibujo arquitectonico': 'Dibujo Arquitectónico',
    'dibujo digital': 'Dibujo Digital',
    'diseño asistido cad': 'Diseño Asistido CAD',
    'tecnología de la información y la comunicación': 'Tecnología de la Información y la Comunicación',
    # Redes 10°
    'administración y soporte a las computadoras': 'Administración y Soporte a las Computadoras',
    'administracion y soporte a las computadoras': 'Administración y Soporte a las Computadoras',
    'administracion y soporte': 'Administración y Soporte a las Computadoras',
    'fundamento de electrotecnia': 'Fundamento de Electrotecnia',
    'fundamentos de programación': 'Fundamentos de Programación',
    'instalaciones eléctricas': 'Instalaciones Eléctricas',
    'instalaciones electricas': 'Instalaciones Eléctricas',
    'instalaciones': 'Instalaciones Eléctricas',
    'tecnologías de información aplicadas a la csrcso': 'Tecnologías de Información Aplicadas a la CSRCSO',
    'english oriented to industrial electrical system': 'English Oriented to Industrial Electrical System',
    'english oriented to': 'English Oriented to Industrial Electrical System',
    # Redes 11°
    'configuración y soporte a redes': 'Configuración y Soporte a Redes',
    'configuracion y soporte a redes': 'Configuración y Soporte a Redes',
    'configuracion y soporte': 'Configuración y Soporte a Redes',
    'mantenimiento de máquinas eléctricas': 'Mantenimiento de Máquinas Eléctricas',
    'mantenimiento': 'Mantenimiento de Máquinas Eléctricas',
    # Redes 12°
    'automatismo industrial': 'Automatismo Industrial',
    "tic\u2019s": 'TICs',
    # Secretariado 10°
    'destrezas digitales': 'Destrezas Digitales',
    'gestión empresarial': 'Gestión Empresarial',
    'english oriented to executive secretary': 'English Oriented to Executive Secretary',
    # Secretariado 11°
    'comunicación empresarial': 'Comunicación Empresarial',
    # Secretariado 12°
    'destrezas computacionales': 'Destrezas Computacionales',
    # Contabilidad 10°
    'contabilidad financiera': 'Contabilidad Financiera',
    'contabilidad': 'Contabilidad Financiera',
    'gestión en tecnologías digitales orientadas a las finanzas': 'Gestión en Tecnologías Digitales Orientadas a las Finanzas',
    'english oriented to accounting and finance': 'English Oriented To Accounting and Finance',
    # Contabilidad 12°
    'gestión en costos': 'Gestión en Costos',
}

SKIP = {'recreo', 'almuerzo', 'guía', 'guia', ''}

def limpiar(texto):
    if not texto: return None
    t = str(texto).strip()
    partes = re.split(r'  +', t)
    nombre = partes[0].strip()
    nombre = re.sub(r'\s+-\s+\w.*$', '', nombre).strip()
    nombre = re.sub(r'\s*-\s*$', '', nombre).strip()
    if len(nombre) <= 2: return None
    clave = nombre.lower().strip()
    return NORM.get(clave, nombre)

# Contar sesiones (celdas no-None) por grupo y materia
grupo_sesiones = defaultdict(lambda: defaultdict(int))
for hoja in GRUPOS:
    ws = wb[hoja]
    for row in ws.iter_rows(min_row=6, values_only=True):
        lec = row[0]
        if lec is None or not isinstance(lec, (int, float)): continue
        for col in MAT_COLS:
            val = row[col-1] if col-1 < len(row) else None
            nombre = limpiar(val)
            if nombre and nombre.lower() not in SKIP:
                grupo_sesiones[hoja][nombre] += 1

# Mostrar conteos
for g in sorted(grupo_sesiones.keys()):
    print(f'=== {g} ===')
    for m, s in sorted(grupo_sesiones[g].items()):
        print(f'  {m}: {s}')
    print()
