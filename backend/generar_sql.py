import openpyxl, re, glob
from collections import defaultdict

ruta = glob.glob(r'C:\Users\Fer\Downloads\HORARIO*.xlsx')[0]
wb = openpyxl.load_workbook(ruta, data_only=True)
GRUPOS = [s for s in wb.sheetnames if re.match(r'\d{2}-\d+', s)]
MAT_COLS_IDX = [2, 4, 6, 8, 10]

NORM = {
    'matemática':'Matemática','matematica':'Matemática','matematicas':'Matemática',
    'español':'Español',
    'inglés':'Inglés','ingles':'Inglés','ingles.':'Inglés',
    'educ. física':'Educación Física','educacion fisica':'Educación Física',
    'educación fisica':'Educación Física',
    'educ. física ana yency garro díaz':'Educación Física',
    'educ. física ana yency garro díaz -  iliana salas guerrero':'Educación Física',
    'est. sociales':'Estudios Sociales','estudios sociales':'Estudios Sociales',
    'estudios socilaes':'Estudios Sociales','estudios soclaes':'Estudios Sociales',
    'cívica':'Cívica','civica':'Cívica',
    'religión':'Religión','religion':'Religión',
    'música':'Música','musica':'Música',
    'física':'Física','fisica':'Física','fisica.':'Física','fisica ':'Física',
    'biología':'Biología','biologia':'Biología',
    'química':'Química','quimica':'Química',
    'psicología':'Psicología','psicologia':'Psicología',
    'emprendedurismo e innovación':'Emprendedurismo',
    'emprendimiento e innovación aplicada a la csrcso':'Emprendedurismo',
    'emprendimiento e innovación aplicada al secretariado ejecutivo':'Emprendedurismo',
    'english for communication':'English For Communication',
    'inglis for comunication':'English For Communication','inglis for':'English For Communication',
    'english for comunnication':'English For Communication',
    'diseño y manufactura asistida por computadora':'Diseño y Manufactura Asistida por Computadora',
    'diseño mano factura asistida':'Diseño y Manufactura Asistida por Computadora',
    'mecanizado con máquinas herramientas':'Mecanizado con Máquinas Herramientas',
    'mecanizado maquinas herramientas':'Mecanizado con Máquinas Herramientas',
    'operaciones de equipo de banco y metrología dimensional':'Operaciones de Equipo de Banco y Metrología Dimensional',
    'tecnología de la información aplicada a la mp':'Tecnología de la Información Aplicada a la MP',
    'dibujo lineal':'Dibujo Lineal',
    'dibujo técnico asistido por computadora':'Dibujo Técnico Asistido por Computadora',
    'dibujo asistido por computadora':'Dibujo Técnico Asistido por Computadora',
    'dibijo mecanico':'Dibujo Mecánico','dibujo mecanico':'Dibujo Mecánico',
    'dibujo arquitectonico':'Dibujo Arquitectónico',
    'dibujo digital':'Dibujo Digital',
    'diseño asistido cad':'Diseño Asistido CAD',
    'tecnología de la información y la comunicación':'Tecnología de la Información y la Comunicación',
    'english oriented to precision mechanics':'English Oriented to Precision Mechanics',
    'administración y soporte a las computadoras':'Administración y Soporte a las Computadoras',
    'administracion y soporte a las computadoras':'Administración y Soporte a las Computadoras',
    'administracion y soporte':'Administración y Soporte a las Computadoras',
    'fundamento de electrotecnia':'Fundamento de Electrotecnia',
    'fundamentos de programación':'Fundamentos de Programación',
    'instalaciones eléctricas':'Instalaciones Eléctricas',
    'instalaciones electricas':'Instalaciones Eléctricas','instalaciones':'Instalaciones Eléctricas',
    'tecnologías de información aplicadas a la csrcso':'Tecnologías de Información Aplicadas a la CSRCSO',
    'english oriented to industrial electrical system':'English Oriented to Industrial Electrical System',
    'english oriented to':'English Oriented to Industrial Electrical System',
    'configuración y soporte a redes':'Configuración y Soporte a Redes',
    'configuracion y soporte a redes':'Configuración y Soporte a Redes',
    'configuracion y soporte':'Configuración y Soporte a Redes',
    'mantenimiento de máquinas eléctricas':'Mantenimiento de Máquinas Eléctricas',
    'mantenimiento':'Mantenimiento de Máquinas Eléctricas',
    'automatismo industrial':'Automatismo Industrial',
    "tic\u2019s":'TICs','tics':'TICs','tic\u00b4s':'TICs',
    'matematicas.':'Matemática',
    'destrezas digitales':'Destrezas Digitales',
    'gestión empresarial':'Gestión Empresarial',
    'comunicación empresarial':'Comunicación Empresarial',
    'english oriented to executive secretary':'English Oriented to Executive Secretary',
    'destrezas computacionales':'Destrezas Computacionales',
    'contabilidad financiera':'Contabilidad Financiera',
    'contabilidad':'Contabilidad Financiera',
    'gestión en tecnologías digitales orientadas a las finanzas':'Gestión en Tecnologías Digitales Orientadas a las Finanzas',
    'english oriented to accounting and finance':'English Oriented To Accounting and Finance',
    'gestión en costos':'Gestión en Costos',
    'gestión en mercadeo':'Gestión en Mercadeo',
}
SKIP = {'recreo','almuerzo','guía','guia',''}

ACADEMICAS = {
    'Matemática','Español','Inglés','Educación Física','Estudios Sociales',
    'Cívica','Religión','Música','Física','Biología','Química','Psicología',
    'Emprendedurismo','English For Communication'
}

def limpiar(texto):
    if not texto: return None
    t = str(texto).strip()
    partes = re.split(r'  +', t)
    nombre = partes[0].strip()
    nombre = re.sub(r'\s+-\s+\w.*$', '', nombre).strip()
    nombre = re.sub(r'\s*-\s*$', '', nombre).strip()
    if len(nombre) <= 2: return None
    clave = nombre.lower().strip()
    return NORM.get(clave, nombre)

def extraer_bloques(ws):
    bloques = []
    esp_actual = ''
    materias_bloque = defaultdict(int)
    en_bloque = False
    for row in ws.iter_rows(values_only=True):
        c0 = str(row[0]).strip() if row[0] else ''
        if 'especialidad' in c0.lower() or ('secci' in c0.lower() and not en_bloque):
            if en_bloque and materias_bloque:
                bloques.append((esp_actual, dict(materias_bloque)))
            esp_actual = c0
            materias_bloque = defaultdict(int)
            en_bloque = True
            continue
        lec = row[0]
        if not isinstance(lec, (int, float)): continue
        for col_idx in MAT_COLS_IDX:
            val = row[col_idx] if col_idx < len(row) else None
            nombre = limpiar(val)
            if nombre and nombre.lower() not in SKIP:
                materias_bloque[nombre] += 1
    if en_bloque and materias_bloque:
        bloques.append((esp_actual, dict(materias_bloque)))
    return bloques

# Recopilar materias con sus niveles y counts máximos
materia_data = {}  # nombre -> {nivel: max_sessions}

for hoja in GRUPOS:
    ws = wb[hoja]
    nivel = int(hoja.split('-')[0])
    bloques = extraer_bloques(ws)
    for esp, mats in bloques:
        for m, cnt in mats.items():
            if m not in materia_data:
                materia_data[m] = {}
            if nivel not in materia_data[m]:
                materia_data[m][nivel] = cnt
            else:
                materia_data[m][nivel] = max(materia_data[m][nivel], cnt)

print("=== RESUMEN DE MATERIAS ENCONTRADAS ===")
for m in sorted(materia_data.keys()):
    print(f"  {m}: {dict(sorted(materia_data[m].items()))}")

# ── Determinar solo_nivel para cada materia ──
def solo_nivel_de(niveles_dict):
    niveles = set(niveles_dict.keys())
    if len(niveles) == 1:
        return list(niveles)[0]
    return None  # aparece en múltiples niveles

# ── Configuración manual de lecciones y espacio ──
# (basado en el análisis del Excel + conocimiento del plan de estudios)
CONFIG = {
    # nombre: (codigo, tipo, lecciones, espacio, bloques_por_sesion, es_tecnica)
    # Académicas
    'Matemática':          ('MAT',  'academica', 5, 'aula_regular',   1, False),
    'Español':             ('ESP',  'academica', 5, 'aula_regular',   1, False),
    'Inglés':              ('ING',  'academica', 3, 'aula_regular',   1, False),
    'Educación Física':    ('EF',   'academica', 3, 'gimnasio',       1, False),
    'Estudios Sociales':   ('SOC',  'academica', 4, 'aula_regular',   1, False),
    'Cívica':              ('CIV',  'academica', 2, 'aula_regular',   1, False),
    'Religión':            ('REL',  'academica', 2, 'aula_regular',   1, False),
    'Música':              ('MUS',  'academica', 2, 'aula_regular',   1, False),
    'Física':              ('FIS',  'academica', 4, 'aula_regular',   1, False),
    'Biología':            ('BIO',  'academica', 4, 'aula_regular',   1, False),
    'Química':             ('QUI',  'academica', 4, 'laboratorio',    1, False),
    'Psicología':          ('PSI',  'academica', 3, 'aula_regular',   1, False),
    'Emprendedurismo':     ('EMP',  'academica', 3, 'aula_regular',   1, False),
    'English For Communication': ('EFC', 'academica', 3, 'aula_regular', 1, False),

    # Mecánica de Precisión 10°
    'Diseño y Manufactura Asistida por Computadora': ('DMAC', 'tecnica', 5, 'taller', 3, True),
    'Mecanizado con Máquinas Herramientas':          ('MMH',  'tecnica', 5, 'taller', 3, True),
    'Operaciones de Equipo de Banco y Metrología Dimensional': ('OEB', 'tecnica', 5, 'taller', 3, True),
    'Tecnología de la Información Aplicada a la MP': ('TIMP', 'tecnica', 3, 'sala_computo', 1, True),
    'Dibujo Lineal':                                 ('DL',   'tecnica', 3, 'taller', 1, True),
    'Dibujo Técnico Asistido por Computadora':       ('DTAC', 'tecnica', 3, 'sala_computo', 1, True),
    'English Oriented to Precision Mechanics':       ('EOPM', 'tecnica', 3, 'aula_regular', 1, True),
    # MP 11°
    'TICs':                                          ('TICS', 'tecnica', 3, 'sala_computo', 1, True),
    'Dibujo Mecánico':                               ('DM',   'tecnica', 3, 'taller', 1, True),
    # MP 12°
    'Dibujo Arquitectónico':                         ('DAR',  'tecnica', 3, 'taller', 1, True),
    'Dibujo Digital':                                ('DD',   'tecnica', 3, 'sala_computo', 1, True),
    'Diseño Asistido CAD':                           ('CAD',  'tecnica', 3, 'sala_computo', 1, True),
    'Tecnología de la Información y la Comunicación':('TIYC', 'tecnica', 3, 'sala_computo', 1, True),

    # Redes/Electrónica 10°
    'Administración y Soporte a las Computadoras':   ('ASC',  'tecnica', 5, 'laboratorio', 3, True),
    'Fundamento de Electrotecnia':                   ('FELE', 'tecnica', 5, 'laboratorio', 3, True),
    'Fundamentos de Programación':                   ('FPRO', 'tecnica', 3, 'sala_computo', 1, True),
    'Instalaciones Eléctricas':                      ('IELE', 'tecnica', 5, 'taller', 3, True),
    'Tecnologías de Información Aplicadas a la CSRCSO': ('TICC', 'tecnica', 3, 'sala_computo', 1, True),
    'English Oriented to Industrial Electrical System': ('EOIES', 'tecnica', 3, 'aula_regular', 1, True),
    # Redes 11°
    'Configuración y Soporte a Redes':               ('CSR',  'tecnica', 5, 'laboratorio', 3, True),
    'Mantenimiento de Máquinas Eléctricas':          ('MME',  'tecnica', 5, 'taller', 3, True),
    # Redes 12°
    'Automatismo Industrial':                        ('AUTO', 'tecnica', 5, 'laboratorio', 3, True),

    # Secretariado 10°
    'Destrezas Digitales':                           ('DDIG', 'tecnica', 3, 'sala_computo', 1, True),
    'Gestión Empresarial':                           ('GEMP', 'tecnica', 5, 'aula_regular', 3, True),
    'English Oriented to Executive Secretary':       ('EOES', 'tecnica', 3, 'aula_regular', 1, True),
    # Secretariado 11°
    'Comunicación Empresarial':                      ('COME', 'tecnica', 5, 'aula_regular', 3, True),
    # Secretariado 12°
    'Destrezas Computacionales':                     ('DCOM', 'tecnica', 3, 'sala_computo', 1, True),

    # Contabilidad 10°
    'Contabilidad Financiera':                       ('CONT', 'tecnica', 5, 'aula_regular', 3, True),
    'Gestión en Tecnologías Digitales Orientadas a las Finanzas': ('GTDF', 'tecnica', 3, 'sala_computo', 1, True),
    'English Oriented To Accounting and Finance':    ('EOAF', 'tecnica', 3, 'aula_regular', 1, True),
    # Contabilidad 11°
    'Gestión en Mercadeo':                           ('GEMC', 'tecnica', 5, 'aula_regular', 3, True),
    # Contabilidad 12°
    'Gestión en Costos':                             ('GCOS', 'tecnica', 5, 'aula_regular', 3, True),
}

# Generar SQL
lines = [
    "-- Generado automáticamente desde HORARIO CTP HEREDIA 2025",
    "-- Insertar solo si no existe (por código)",
    "",
]

for nombre, niveles in sorted(materia_data.items()):
    cfg = CONFIG.get(nombre)
    if not cfg:
        lines.append(f"-- ADVERTENCIA: sin config para: {nombre}")
        continue

    codigo, tipo, lecciones, espacio, bloques, es_tecnica = cfg
    solo_nivel = solo_nivel_de(niveles)

    # Para académicas que aparecen en múltiples niveles, no restringir
    if nombre in ACADEMICAS:
        solo_nivel = None

    solo_nivel_sql = str(solo_nivel) if solo_nivel else 'NULL'
    es_tecnica_sql = 'TRUE' if es_tecnica else 'FALSE'
    nombre_escaped = nombre.replace("'", "''")

    lines.append(
        f"INSERT INTO materias (nombre, codigo, tipo, lecciones_semanales, "
        f"requiere_espacio, bloques_por_sesion, es_tecnica, solo_nivel, activo, dias_permitidos) "
        f"VALUES ('{nombre_escaped}', '{codigo}', '{tipo}', {lecciones}, "
        f"'{espacio}', {bloques}, {es_tecnica_sql}, {solo_nivel_sql}, TRUE, ARRAY[]::text[])"
        f" ON CONFLICT (codigo) DO NOTHING;"
    )

sql_content = '\n'.join(lines)

with open('insert_materias.sql', 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"\n=== SQL generado: insert_materias.sql ===")
print(f"Total materias configuradas: {sum(1 for n in materia_data if CONFIG.get(n))}")
print(f"Sin configuración: {[n for n in materia_data if not CONFIG.get(n)]}")
print("\nContenido:")
print(sql_content)
